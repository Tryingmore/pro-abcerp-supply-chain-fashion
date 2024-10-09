# -*- coding: utf-8 -*-
# import sys
# import os
import openpyxl
import ast
import os
import re
import zipfile,datetime,time
import concurrent.futures
import subprocess
# import MySQLdb
# import MySQLdb.cursors
# import xlrd

def read_xls():
    wb = openpyxl.load_workbook('/data/221128OCA合并模块119个-1202.xlsx',data_only=True)
    names = wb.sheetnames

    sheet = wb[names[0]]

    maxRow = sheet.max_row

    maxColumn = sheet.max_column

    current_sheet = wb.active

    current_name = sheet.title

    n = 2
    path_c = []
    res = []
    while n <= maxRow:

        path1 = sheet.cell(n, 10).value
        path2 = sheet.cell(n, 1).value
        path_dir = sheet.cell(n, 7).value
        path_all = path_dir

        if not path1:
            res = res + [(path2,path_all)]
        n = n + 1
    return res

def execute_command_with_no_out(cmd, timeout):
    p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE,stderr=subprocess.PIPE, close_fds=True)

    try:
        p.wait(timeout=timeout)
        if p.poll:

            outs, errs = p.communicate()
            return outs, errs
    except subprocess.TimeoutExpired as e:
        p.kill()

def run_git(git_commit,repo):
    time.sleep(1)
    com = execute_command_with_no_out(git_commit, 10)
    if com[1]:
        print(repo,git_commit,'False',com[1])
    else:
        print(repo,git_commit,'sussuss',com[0])

if __name__ == "__main__":

    time_1 = time.time()

    time_2 = time.time()
    repos = read_xls()
    counts = len(repos)
    i = 1
    for l in repos[0:1]:
        repo = l[0]
        source_path = l[1] + repo
        to_path = l[1].replace('OCA/OCA16','odoo16/cn_oca')
        print(i,'/',counts,repo,source_path,to_path)
        i +=1
        g1 = "git checkout 16.0 > /dev/null 2>&1"
        g2 = "git branch -v > /dev/null 2>&1"
        g3 = 'git checkout -b oca_mig__%s > /dev/null 2>&1' %(repo)
        g31 = 'mkdir -p %s' %(to_path)
        g4 = 'cp -R %s  %s' %(source_path,to_path)
        g5 = 'git add .'
        g6 = 'git commit -m"%s" ' %(repo)
        g7 = 'git push --set-upstream origin oca_mig__%s' %(repo)
        com = run_git('cd /data/odoo16/cn_oca',repo)
        com = run_git(g1, repo)
        com = run_git(g2, repo)
        com = run_git(g3, repo)
        com = run_git(g31, repo)
        com = run_git(g4, repo)
        com = run_git('cd /data/odoo16/cn_oca',repo)
        com = run_git(g5, repo)
        com = run_git(g6, repo)
        com = run_git(g7, repo)
    use_time = time_2 - time_1
    # print(f'use time:{use_time}')
