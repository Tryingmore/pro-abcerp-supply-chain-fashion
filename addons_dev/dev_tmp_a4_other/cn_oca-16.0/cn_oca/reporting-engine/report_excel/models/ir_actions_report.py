# -*- coding: utf-8 -*-

from copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
from contextlib import closing
from tempfile import NamedTemporaryFile
import odoo
from odoo import api, models, fields, _
from odoo.exceptions import AccessDenied, UserError, ValidationError
from odoo.tools.safe_eval import safe_eval, wrap_module
from odoo.tools import html2plaintext

import logging, subprocess, tempfile, os, base64, math

_logger = logging.getLogger(__name__)


class IrActionsReport(models.Model):
    _inherit = "ir.actions.report"

    report_type = fields.Selection(selection_add=[('excel', 'Excel')], ondelete={"excel": "set default"})
    template_path = fields.Char('Template file path', help='Excel temnplate file path, eg: /report_excel/data/template.xlsx')

    @api.model
    def render_excel(self, docids, data=None):
        if not data:
            data = {}
        data.setdefault('report_type', 'text')
        data = self._get_rendering_context(self, docids, data)

        eval_args = {'self': self, 'data': data, 'datetime.datetime': datetime, 'datetime.timedelta': timedelta, 'sorted': sorted, 'html2plaintext': html2plaintext, 'math': wrap_module(__import__('math'), [])}

        def _copy_cell_style(cell, target_cell):
            target_cell.style = copy(cell.style)
            target_cell.font = copy(cell.font)
            target_cell.border = copy(cell.border)
            target_cell.fill = copy(cell.fill)
            target_cell.number_format = copy(cell.number_format)
            target_cell.protection = copy(cell.protection)
            target_cell.alignment = copy(cell.alignment)
            # 复制单元格合并信息
            merged_list = []
            for range in cell.parent.merged_cells.ranges:
                if cell.coordinate in range and range.start_cell == cell:
                    merged_list.append([target_cell.row, target_cell.column, target_cell.row + range.size['rows']-1, target_cell.column + range.size['columns']-1])
            for item in merged_list:
                cell.parent.merge_cells(start_row=item[0], start_column=item[1], end_row=item[2], end_column=item[3])

        # pyopenxl正常的插入行方法不会处理下方的合并单元格，会导致显示错位，需要在合并后重新合并单元格
        def _insert_rows(sheet, row, count):
            merged_list = []
            for range in sheet.merged_cells.ranges:
                if range.min_row > row:
                    sheet.unmerge_cells(start_row=range.min_row, start_column=range.min_col, end_row=range.max_row, end_column=range.max_col)
                    merged_list.append([range.min_row + count, range.min_col, range.max_row + count, range.max_col])
            sheet.insert_rows(row, count)
            for item in merged_list:
                sheet.merge_cells(start_row=item[0], start_column=item[1], end_row=item[2], end_column=item[3])

        def render_work_sheet(ws, docs, doc=None):
            # 常规内容
            for r in range(ws.max_row):
                for c in range(ws.max_column):
                    value = ws.cell(row=r + 1, column=c + 1).value
                    if value and isinstance(value, str) and value.startswith('[') and value.endswith(']'):
                        try:
                            value = value[1:-1]
                            ws.cell(row=r + 1, column=c + 1).value = safe_eval(value, {**eval_args, **{'docs': docs, 'doc': doc}})
                        except Exception as ex:
                            ws.cell(row=r + 1, column=c + 1).value = str(ex)

            # 处理循环，从注释中读取，格式为 doc.invoice_ids:inv
            # 先插入行，解决一个sheet中有多个表第一个表插入行数过多，下面的表不出来的问题
            for r in range(ws.max_row):
                comment = ws.cell(row=r + 1, column=1).comment
                if comment and comment.text and '|' in comment.text:
                    list, item_str = comment.text.split('|')
                    list = safe_eval(list, {**eval_args, **{'docs': docs, 'doc': doc}})
                    if len(list) > 1:
                        _insert_rows(ws, r + 2, len(list) - 1)        # 插入行
                    elif len(list) == 0:
                        ws.delete_rows(r + 1, 1)
            # 写入数据
            for r in range(ws.max_row):
                comment = ws.cell(row=r + 1, column=1).comment
                if comment and comment.text and '|' in comment.text:
                    list, item_str = comment.text.split('|')
                    list = safe_eval(list, {**eval_args, **{'docs': docs, 'doc': doc}})
                    for c in range(ws.max_column):                  # 按列循环写入内容
                        value, i = ws.cell(row=r + 1, column=c + 1).value, 0
                        if value and isinstance(value, str) and value.startswith('<') and value.endswith('>'):
                            value = value[1:-1]
                            for item in list:
                                if i > 0:
                                    _copy_cell_style(ws.cell(row=r + 1, column=c + 1), ws.cell(row=r + i + 1, column=c + 1))     # 复制单元格格式
                                try:
                                    ws.cell(row=r + i + 1, column=c + 1).value = safe_eval(value, {**eval_args, **{'docs': docs, 'doc': doc, item_str: item, 'index': i}})
                                except Exception as ex:
                                    try:                            # 合并单元格只能往第一个格子里面写入
                                        ws.cell(row=r + i + 1, column=c + 1).value = str(ex)
                                    except Exception as ex1:        # pragma: no cover
                                        _logger.error(ex1)
                                i += 1
                        if not value and len(list) > 1:
                            for i in range(len(list)):
                                if i > 0:
                                    _copy_cell_style(ws.cell(row=r + 1, column=c + 1), ws.cell(row=r + i + 1, column=c + 1))     # 复制单元格格式
                    ws.cell(row=r + 1, column=1).comment = None

        def _format_worksheet_name(name):
            for char in ['[', ']', '*', '/', '\\', '?', ':']:
                name = name.replace(char, '_')
            return name

        filepath = self._get_full_excel_template_path()
        wb = load_workbook(filepath)
        if not self.multi and len(docids) > 1:
            raise ValidationError('当前报表的设置不支持同时输出多个单据！')
        defined_names = wb.defined_names
        defined_names = dict([(name, defined_names[name].value.replace('"','')) for name in defined_names])
        docs_split_by = defined_names.get('DOCS_SPLIT_BY', 'LINE')
        sheet_name_role = defined_names.get('SHEET_NAME_ROLE', 'doc.display_name')
        # 数据准备
        model = self.model
        if data.get('docids', False) and data.get('res_model', False):
            model = data.get('res_model', False)
            docids = data.get('docids', [])
        docs = self.env[model].browse(docids)
        # 按工作表拆分的情况，先复制工作表
        if docs_split_by == 'SHEET':
            if len(wb.worksheets) > 1:
                raise ValidationError('多工作表的模板不支持按工作表分隔多个单据！')
            ws = wb.worksheets[0]
            for i in range(len(docs)):
                if i > 0:
                    ws_copy = wb.copy_worksheet(ws)                                 # 先复制工作表
                    for img in ws._images: ws_copy.add_image(copy(img))             # copy_worksheet不复制图片，单独复制一下;
                    # openpyxl复制图片有个BUG，在 https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1843 报了个issue，可以参考一下改源码def _data方法修复
            for i in range(len(docs)):
                ws = wb.worksheets[i]
                ws.title = _format_worksheet_name(safe_eval(sheet_name_role, {**eval_args, **{'docs': docs, 'doc': docs[i]}}))
                render_work_sheet(ws, docs, docs[i])
        else:
            # 根据模板文件渲染文件内容
            for ws in wb:
                render_work_sheet(ws, docs, docs[0])
        # 输出文件
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream, 'excel'

    def _get_full_excel_template_path(self):
        if self.template_path:
            for path in odoo.addons.__path__:
                if os.path.exists(path + self.template_path):
                    return path + self.template_path
        return False

    @api.constrains('template_path')
    def _check_template_path(self):
        template_path = self._get_full_excel_template_path()
        if not template_path:
            raise ValidationError('The template file does not exist, please check the file path!\n' + self.template_path)
